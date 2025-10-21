"""
exporter/freee_exporter.py - freee用Excel出力クラス
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class BaseExporter:
    """Excel出力の基底クラス"""
    
    def __init__(self, output_dir=None):
        if output_dir is None:
            self.output_dir = Path.home() / "Desktop"
        else:
            self.output_dir = Path(output_dir)
    
    def _generate_filename(self, original_filename, suffix=""):
        """出力ファイル名を生成する（重複時は連番を付与）"""
        # 元のファイル名から拡張子を除去
        base_name = Path(original_filename).stem
        
        # 現在の日付を取得
        today = datetime.now().strftime('%Y%m%d')
        
        # 基本のファイル名
        if suffix:
            filename = f"{base_name}_{today}{suffix}.xlsx"
        else:
            filename = f"{base_name}_{today}.xlsx"
        
        # 重複チェックと連番付与
        counter = 1
        while (self.output_dir / filename).exists():
            if suffix:
                filename = f"{base_name}_{today}{suffix}_{counter:02d}.xlsx"
            else:
                filename = f"{base_name}_{today}_{counter:02d}.xlsx"
            counter += 1
        
        return filename


class TestExcelExporter(BaseExporter):
    """テスト用Excel出力クラス"""
    
    def export(self, data_list: list[dict], original_filename: str) -> str:
        """テスト用のExcelファイルに出力する
        
        Args:
            data_list: list[dict] - 検証済みデータリスト
            original_filename: str - 元のファイル名
        
        Returns:
            str: 出力ファイルパス
        """
        # ファイル名を生成
        output_filename = self._generate_filename(original_filename)
        output_path = self.output_dir / output_filename
        
        # 出力用データを作成
        output_data = []
        error_rows = []
        
        for idx, data in enumerate(data_list):
            row_data = {
                'シート名': data.get('シート名', ''),
                '日付': data.get('日付', ''),
                '金額': data.get('金額', '')
            }
            output_data.append(row_data)
            
            # エラーがある行を記録
            errors = data.get('_errors', [])
            if errors:
                error_rows.append(idx + 2)  # +2 はヘッダー行を考慮
        
        # DataFrameに変換してExcel出力
        df = pd.DataFrame(output_data)
        df.to_excel(output_path, index=False, sheet_name='test_data', engine='openpyxl')
        
        # エラー行を赤色にする
        if error_rows:
            self._highlight_error_rows(output_path, 'test_data', error_rows, len(df.columns))
        
        return str(output_path)
    
    def _highlight_error_rows(self, file_path, sheet_name, error_rows, num_columns):
        """エラー行を赤色でハイライト
        
        Args:
            file_path: str - Excelファイルパス
            sheet_name: str - シート名
            error_rows: list[int] - エラー行番号のリスト
            num_columns: int - 列数
        """
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        
        # 薄い赤（ピンク）の塗りつぶし
        pink_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        
        for row_num in error_rows:
            for col_num in range(1, num_columns + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = pink_fill
        
        wb.save(file_path)


class FreeeExcelExporter(BaseExporter):
    """freee用Excel出力クラス"""
    
    # freee用の列名定義
    FREEE_COLUMNS = [
        '[表題行]', '日付', '伝票番号', '決算整理仕訳',
        '借方勘定科目', '借方科目コード', '借方補助科目', '借方取引先', '借方取引先コード',
        '借方部門', '借方品目', '借方メモタグ', '借方セグメント1', '借方セグメント2', '借方セグメント3',
        '借方金額', '借方税区分', '借方税額',
        '貸方勘定科目', '貸方科目コード', '貸方補助科目', '貸方取引先', '貸方取引先コード',
        '貸方部門', '貸方品目', '貸方メモタグ', '貸方セグメント1', '貸方セグメント2', '貸方セグメント3',
        '貸方金額', '貸方税区分', '貸方税額', '摘要'
    ]
    
    def export(self, data_list: list[dict], original_filename: str) -> str:
        """freee用のExcelファイルに出力する
        
        Args:
            data_list: list[dict] - 検証済みデータリスト
            original_filename: str - 元のファイル名
        
        Returns:
            str: 出力ファイルパス
        """
        # ファイル名を生成
        output_filename = self._generate_filename(original_filename)
        output_path = self.output_dir / output_filename
        
        # freee用のデータに変換
        freee_data = []
        error_rows = []
        color_info = []  # 色付け情報 {'row': 行番号, 'col': 列名, 'color': 色}
        
        for idx, data in enumerate(data_list):
            row = {}
            
            # freee用の全列について処理
            for col in self.FREEE_COLUMNS:
                if col in data:
                    row[col] = data[col]
                elif col == '日付':
                    row[col] = data.get('日付', '')
                elif col == '借方金額' or col == '貸方金額':
                    row[col] = data.get('金額', '')
                else:
                    row[col] = ''
            
            # 候補列を追加
            row['候補'] = data.get('候補', '')
            
            # 借方取引先の色付け情報
            borrow_match = data.get('借方取引先_match_type', 'none')
            if borrow_match == 'partner_list' or borrow_match == 'freee_exact':
                color_info.append({'row': idx + 2, 'col': '借方取引先', 'color': 'green'})
            elif borrow_match == 'fuzzy':
                color_info.append({'row': idx + 2, 'col': '借方取引先', 'color': 'red'})
            
            # 貸方取引先の色付け情報
            lend_match = data.get('貸方取引先_match_type', 'none')
            if lend_match == 'partner_list' or lend_match == 'freee_exact':
                color_info.append({'row': idx + 2, 'col': '貸方取引先', 'color': 'green'})
            elif lend_match == 'fuzzy':
                color_info.append({'row': idx + 2, 'col': '貸方取引先', 'color': 'red'})
            
            # エラーがある場合
            errors = data.get('_errors', [])
            if errors:
                error_rows.append(idx + 2)  # +2 はヘッダー行を考慮
                # エラーメッセージを結合
                if isinstance(errors, list):
                    row['エラー内容'] = '\n'.join(errors)
                else:
                    row['エラー内容'] = str(errors)
            else:
                row['エラー内容'] = ''
            
            freee_data.append(row)
        
        # 出力列: freee列 + 候補 + エラー内容
        output_columns = self.FREEE_COLUMNS + ['候補', 'エラー内容']
        
        # DataFrameに変換してExcel出力
        df = pd.DataFrame(freee_data, columns=output_columns)
        df.to_excel(output_path, index=False, sheet_name='freee_data', engine='openpyxl')
        
        # エラー行を赤色にする
        if error_rows:
            self._highlight_error_rows(output_path, 'freee_data', error_rows, len(output_columns))
        
        # 取引先セルを色付け
        if color_info:
            self._color_partner_cells(output_path, 'freee_data', color_info, output_columns)
        
        # 列幅を自動調整
        self._adjust_column_widths(output_path, 'freee_data', df)
        
        return str(output_path)
    
    def _highlight_error_rows(self, file_path, sheet_name, error_rows, num_columns):
        """エラー行を赤色でハイライト
        
        Args:
            file_path: str - Excelファイルパス
            sheet_name: str - シート名
            error_rows: list[int] - エラー行番号のリスト
            num_columns: int - 列数
        """
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        
        # 薄い赤（ピンク）の塗りつぶし
        pink_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        
        for row_num in error_rows:
            for col_num in range(1, num_columns + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = pink_fill
        
        wb.save(file_path)
    
    def _color_partner_cells(self, file_path, sheet_name, color_info, columns):
        """取引先セルを色付け
        
        Args:
            file_path: str - Excelファイルパス
            sheet_name: str - シート名
            color_info: list[dict] - 色付け情報
            columns: list - 列名リスト
        """
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        
        # 緑色の塗りつぶし
        green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        # 薄い赤（ピンク）の塗りつぶし
        pink_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        
        # 列名から列番号へのマッピング
        col_mapping = {col: idx + 1 for idx, col in enumerate(columns)}
        
        for info in color_info:
            row_num = info['row']
            col_name = info['col']
            color = info['color']
            
            if col_name in col_mapping:
                col_num = col_mapping[col_name]
                cell = ws.cell(row=row_num, column=col_num)
                
                if color == 'green':
                    cell.fill = green_fill
                elif color == 'red':
                    cell.fill = pink_fill
        
        wb.save(file_path)
    
    def _adjust_column_widths(self, file_path, sheet_name, df):
        """列幅を自動調整する
        
        Args:
            file_path: str - Excelファイルパス
            sheet_name: str - シート名
            df: DataFrame - データフレーム
        """
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        
        for col_idx, column in enumerate(df.columns, start=1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            
            # 列のデータをチェック（ヘッダー以外）
            has_data = any(
                df.iloc[row_idx, col_idx - 1] not in ['', None, 'nan']
                for row_idx in range(len(df))
            )
            
            if not has_data:
                # データがない列は幅を小さく
                ws.column_dimensions[column_letter].width = 8
            else:
                # データがある列は自動調整
                max_length = len(str(column))  # ヘッダーの長さ
                
                for row_idx in range(len(df)):
                    cell_value = str(df.iloc[row_idx, col_idx - 1])
                    if cell_value and cell_value != 'nan':
                        # 日本語文字は2文字分としてカウント
                        length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                        max_length = max(max_length, length)
                
                # 最大幅を50に制限
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        wb.save(file_path)
        